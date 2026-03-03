import glob
import re
import sqlite3
import unicodedata
import zipfile
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


NS = {"kml": "http://www.opengis.net/kml/2.2"}


@dataclass
class ImportStats:
    source: str
    read: int = 0
    inserted: int = 0
    skipped_no_coord: int = 0
    dedup_coord: int = 0
    enriched: int = 0


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def normalize_slug(text):
    if not text:
        return "sem-bairro"
    raw = unicodedata.normalize("NFKD", text)
    raw = raw.encode("ascii", "ignore").decode("ascii")
    raw = re.sub(r"[^a-zA-Z0-9]+", "-", raw).strip("-").lower()
    return raw or "sem-bairro"


def parse_float(value):
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def coord_key(lat, lng):
    if lat is None or lng is None:
        return None
    return f"{lat:.7f},{lng:.7f}"


def parse_lamp_power_from_ativo(ativo):
    if not ativo:
        return (None, None)
    # Ex.: B3-L150 x1, L060 x1
    m = re.search(r"L(\d{2,3})", ativo.upper())
    power = int(m.group(1)) if m else None
    return (ativo.strip(), power)


def load_existing_coords(cur):
    cur.execute("SELECT id, lat, lng FROM pontos_ilp WHERE lat IS NOT NULL AND lng IS NOT NULL")
    idx = {}
    for row_id, lat, lng in cur.fetchall():
        key = coord_key(lat, lng)
        if key:
            idx[key] = row_id
    return idx


def enrich_existing(cur, point_id, bairro=None, tipo_lampada=None, potencia=None):
    cur.execute(
        """
        UPDATE pontos_ilp
        SET
          bairro = COALESCE(bairro, ?),
          tipo_lampada = COALESCE(tipo_lampada, ?),
          potencia = COALESCE(potencia, ?)
        WHERE id = ?
        """,
        (bairro, tipo_lampada, potencia, point_id),
    )
    return cur.rowcount > 0


def insert_point(cur, payload):
    cur.execute(
        """
        INSERT OR IGNORE INTO pontos_ilp
          (etiqueta, endereco, bairro, cidade, lat, lng, tipo_poste, altura, tipo_luminaria, braco, tipo_lampada, potencia, status)
        VALUES
          (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            payload["etiqueta"],
            payload.get("endereco"),
            payload.get("bairro"),
            payload.get("cidade", "Macapa"),
            payload.get("lat"),
            payload.get("lng"),
            payload.get("tipo_poste"),
            payload.get("altura"),
            payload.get("tipo_luminaria"),
            payload.get("braco"),
            payload.get("tipo_lampada"),
            payload.get("potencia"),
            payload.get("status", "cadastrado"),
        ),
    )
    return cur.rowcount > 0


def import_xlsx_base(cur, coord_idx, xlsx_path):
    stats = ImportStats("BASE 09 - 2025.xlsx")
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats.read += 1
        barramento = normalize_text(row[0])
        tipo = normalize_text(row[1])
        potencia = row[2]
        lat = parse_float(row[7])
        lng = parse_float(row[8])
        if lat is None or lng is None:
            stats.skipped_no_coord += 1
            continue

        key = coord_key(lat, lng)
        if key in coord_idx:
            if enrich_existing(cur, coord_idx[key], tipo_lampada=tipo, potencia=int(potencia) if potencia is not None else None):
                stats.enriched += 1
            stats.dedup_coord += 1
            continue

        etiqueta = f"BASE09-{barramento}" if barramento else f"BASE09-PT-{stats.read:06d}"
        payload = {
            "etiqueta": etiqueta,
            "endereco": "Sem endereco (BASE 09)",
            "bairro": None,
            "cidade": "Macapa",
            "lat": lat,
            "lng": lng,
            "tipo_lampada": tipo,
            "potencia": int(potencia) if potencia is not None else None,
            "status": "cadastrado",
        }
        if insert_point(cur, payload):
            cur.execute("SELECT id FROM pontos_ilp WHERE etiqueta = ?", (etiqueta,))
            point_id = cur.fetchone()[0]
            coord_idx[key] = point_id
            stats.inserted += 1

    return stats


def parse_kmz_points(kmz_path):
    with zipfile.ZipFile(kmz_path, "r") as zf:
        kml_name = [n for n in zf.namelist() if n.lower().endswith(".kml")][0]
        root = ET.fromstring(zf.read(kml_name))
    for pm in root.findall(".//kml:Placemark", NS):
        coord_node = pm.find(".//kml:Point/kml:coordinates", NS)
        if coord_node is None or not (coord_node.text or "").strip():
            continue
        coords = (coord_node.text or "").strip().split(",")
        if len(coords) < 2:
            continue
        lng = parse_float(coords[0])
        lat = parse_float(coords[1])
        if lat is None or lng is None:
            continue
        simple = {}
        for sd in pm.findall(".//kml:SimpleData", NS):
            simple[sd.attrib.get("name")] = (sd.text or "").strip()
        yield (lat, lng, simple)


def import_kmz_selt(cur, coord_idx, kmz_path):
    stats = ImportStats(Path(kmz_path).name)
    seq = 0
    for lat, lng, simple in parse_kmz_points(kmz_path):
        stats.read += 1
        key = coord_key(lat, lng)
        etiqueta_raw = normalize_text(simple.get("ETIQUETA"))
        ativo = normalize_text(simple.get("ATIVO"))
        tipo_lampada, potencia = parse_lamp_power_from_ativo(ativo)

        if key in coord_idx:
            if enrich_existing(cur, coord_idx[key], tipo_lampada=tipo_lampada, potencia=potencia):
                stats.enriched += 1
            stats.dedup_coord += 1
            continue

        seq += 1
        if etiqueta_raw:
            etiqueta = f"SELT-{etiqueta_raw.lstrip('*')}"
        else:
            etiqueta = f"SELT-PT-{seq:06d}"

        payload = {
            "etiqueta": etiqueta,
            "endereco": "Sem endereco (SELT)",
            "bairro": None,
            "cidade": "Macapa",
            "lat": lat,
            "lng": lng,
            "tipo_lampada": tipo_lampada,
            "potencia": potencia,
            "status": "cadastrado",
        }
        if insert_point(cur, payload):
            cur.execute("SELECT id FROM pontos_ilp WHERE etiqueta = ?", (etiqueta,))
            point_id = cur.fetchone()[0]
            coord_idx[key] = point_id
            stats.inserted += 1
    return stats


def extract_bairro_from_filename(name):
    m = re.search(r"BAIRRO\s+(.+?)(?:\s+-\s+\d+a\s+ET|\s+-\s+2a\s+ETAPA|\s*\.kmz|$)", name, flags=re.IGNORECASE)
    if m:
        return normalize_text(m.group(1).replace("KMZ", ""))
    # fallback para nomes simples
    m2 = re.search(r"BAIRRO\s+(.+?)(?:\.kmz|$)", name, flags=re.IGNORECASE)
    return normalize_text(m2.group(1)) if m2 else None


def import_kmz_bairros(cur, coord_idx, kmz_paths):
    all_stats = []
    for kmz_path in kmz_paths:
        name = Path(kmz_path).name
        stats = ImportStats(name)
        bairro = extract_bairro_from_filename(name)
        slug = normalize_slug(bairro or name)
        seq = 0

        for lat, lng, _simple in parse_kmz_points(kmz_path):
            stats.read += 1
            key = coord_key(lat, lng)
            if key in coord_idx:
                if bairro and enrich_existing(cur, coord_idx[key], bairro=bairro):
                    stats.enriched += 1
                stats.dedup_coord += 1
                continue

            seq += 1
            etiqueta = f"KMZ-{slug}-{seq:06d}"
            payload = {
                "etiqueta": etiqueta,
                "endereco": f"Sem endereco ({bairro or 'Bairro KMZ'})",
                "bairro": bairro,
                "cidade": "Macapa",
                "lat": lat,
                "lng": lng,
                "status": "cadastrado",
            }
            if insert_point(cur, payload):
                cur.execute("SELECT id FROM pontos_ilp WHERE etiqueta = ?", (etiqueta,))
                point_id = cur.fetchone()[0]
                coord_idx[key] = point_id
                stats.inserted += 1

        all_stats.append(stats)
    return all_stats


def main():
    repo = Path(__file__).resolve().parents[1]
    db_path = repo / "macapaluz.db"
    downloads = Path(r"C:\Users\Super economico\Downloads")
    xlsx_path = downloads / "BASE 09 - 2025.xlsx"
    kmz_selt = glob.glob(str(downloads / "*SELT*.kmz"))
    kmz_all = glob.glob(str(downloads / "*.kmz"))

    if not db_path.exists():
        raise FileNotFoundError(f"Banco nao encontrado: {db_path}")
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Planilha nao encontrada: {xlsx_path}")
    if not kmz_selt:
        raise FileNotFoundError("Arquivo KMZ SELT nao encontrado em Downloads.")

    kmz_selt_path = kmz_selt[0]
    kmz_bairros = [p for p in kmz_all if str(p).lower() != str(kmz_selt_path).lower()]

    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute("PRAGMA foreign_keys = ON;")

    coord_idx = load_existing_coords(cur)
    stats = []
    stats.append(import_xlsx_base(cur, coord_idx, str(xlsx_path)))
    stats.append(import_kmz_selt(cur, coord_idx, kmz_selt_path))
    stats.extend(import_kmz_bairros(cur, coord_idx, kmz_bairros))

    conn.commit()

    cur.execute("SELECT COUNT(*) FROM pontos_ilp")
    total_points = cur.fetchone()[0]
    conn.close()

    print("IMPORTACAO CONCLUIDA")
    for s in stats:
        print(
            f"- {s.source}: lidos={s.read} inseridos={s.inserted} "
            f"dedup_coord={s.dedup_coord} enriquecidos={s.enriched} sem_coord={s.skipped_no_coord}"
        )
    print(f"TOTAL pontos_ilp: {total_points}")


if __name__ == "__main__":
    main()
