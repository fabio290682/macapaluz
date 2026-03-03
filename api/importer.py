import csv
import io
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


NS = {"kml": "http://www.opengis.net/kml/2.2"}


def normalize_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text if text else None


def parse_float(value):
    if value is None:
        return None
    text = str(value).strip().replace(",", ".")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def parse_int(value):
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    text = re.sub(r"[^\d-]", "", text)
    try:
        return int(text)
    except ValueError:
        return None


def _coord_key(lat, lng):
    if lat is None or lng is None:
        return None
    return f"{lat:.7f},{lng:.7f}"


def _first_value(row, keys):
    for key in keys:
        if key in row and normalize_text(row[key]) is not None:
            return normalize_text(row[key])
    return None


def _parse_csv(content):
    text = content.decode("utf-8", errors="replace")
    sample = text[:4096]
    delimiter = ";" if sample.count(";") > sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    points = []
    for row in reader:
        row_l = {str(k).strip().lower(): v for k, v in row.items() if k is not None}
        lat = parse_float(_first_value(row_l, ["lat", "latitude", "y"]))
        lng = parse_float(_first_value(row_l, ["lng", "lon", "longitude", "x"]))
        points.append(
            {
                "etiqueta": _first_value(row_l, ["etiqueta", "id", "codigo", "ponto"]),
                "endereco": _first_value(row_l, ["endereco", "logradouro", "rua"]) or "Sem endereco",
                "bairro": _first_value(row_l, ["bairro"]),
                "cidade": _first_value(row_l, ["cidade"]) or "Macapa",
                "lat": lat,
                "lng": lng,
                "tipo_lampada": _first_value(row_l, ["tipo_lampada", "lampada", "tipo"]),
                "potencia": parse_int(_first_value(row_l, ["potencia", "pot", "watt"])),
                "status": _first_value(row_l, ["status"]) or "cadastrado",
            }
        )
    return points


def _extract_kml_points(kml_bytes):
    root = ET.fromstring(kml_bytes)
    points = []
    for pm in root.findall(".//kml:Placemark", NS):
        coord_node = pm.find(".//kml:Point/kml:coordinates", NS)
        if coord_node is None:
            continue
        coords = (coord_node.text or "").strip().split(",")
        if len(coords) < 2:
            continue
        lng = parse_float(coords[0])
        lat = parse_float(coords[1])
        if lat is None or lng is None:
            continue

        attrs = {}
        for sd in pm.findall(".//kml:SimpleData", NS):
            name = (sd.attrib.get("name") or "").strip().lower()
            attrs[name] = normalize_text(sd.text)

        name_node = pm.find("./kml:name", NS)
        etiqueta = normalize_text(name_node.text if name_node is not None else None) or attrs.get("etiqueta")
        points.append(
            {
                "etiqueta": etiqueta,
                "endereco": attrs.get("endereco") or "Sem endereco",
                "bairro": attrs.get("bairro"),
                "cidade": attrs.get("cidade") or "Macapa",
                "lat": lat,
                "lng": lng,
                "tipo_lampada": attrs.get("tipo_lampada") or attrs.get("ativo"),
                "potencia": parse_int(attrs.get("potencia")),
                "status": attrs.get("status") or "cadastrado",
            }
        )
    return points


def _parse_kml(content):
    return _extract_kml_points(content)


def _parse_kmz(content):
    with zipfile.ZipFile(io.BytesIO(content), "r") as zf:
        names = [n for n in zf.namelist() if n.lower().endswith(".kml")]
        if not names:
            return []
        return _extract_kml_points(zf.read(names[0]))


def _parse_xlsx(content):
    if load_workbook is None:
        raise ValueError("Dependencia openpyxl indisponivel para XLSX")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=1, values_only=True)
    headers = [str(h).strip().lower() if h is not None else "" for h in next(rows, [])]
    points = []
    for row in rows:
        row_d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        lat = parse_float(_first_value(row_d, ["lat", "latitude", "y"]))
        lng = parse_float(_first_value(row_d, ["lng", "lon", "longitude", "x"]))
        points.append(
            {
                "etiqueta": _first_value(row_d, ["etiqueta", "id", "codigo", "ponto"]),
                "endereco": _first_value(row_d, ["endereco", "logradouro", "rua"]) or "Sem endereco",
                "bairro": _first_value(row_d, ["bairro"]),
                "cidade": _first_value(row_d, ["cidade"]) or "Macapa",
                "lat": lat,
                "lng": lng,
                "tipo_lampada": _first_value(row_d, ["tipo_lampada", "lampada", "tipo"]),
                "potencia": parse_int(_first_value(row_d, ["potencia", "pot", "watt"])),
                "status": _first_value(row_d, ["status"]) or "cadastrado",
            }
        )
    return points


def parse_uploaded_file(filename, content):
    ext = Path(filename or "").suffix.lower()
    if ext == ".csv":
        return _parse_csv(content)
    if ext == ".kml":
        return _parse_kml(content)
    if ext == ".kmz":
        return _parse_kmz(content)
    if ext == ".xlsx":
        return _parse_xlsx(content)
    raise ValueError(f"Formato nao suportado: {ext}")


def import_points_to_db(conn, points, origem_dado="upload"):
    cur = conn.cursor()
    cur.execute("SELECT id, lat, lng FROM pontos_ilp WHERE lat IS NOT NULL AND lng IS NOT NULL")
    coord_idx = {}
    for row in cur.fetchall():
        key = _coord_key(row[1], row[2])
        if key:
            coord_idx[key] = row[0]

    cur.execute("SELECT etiqueta FROM pontos_ilp")
    existing_tags = {r[0] for r in cur.fetchall() if r[0]}

    stats = {"read": len(points), "inserted": 0, "updated": 0, "skipped_no_coord": 0, "dedup_coord": 0}
    seq = 0

    for p in points:
        lat = p.get("lat")
        lng = p.get("lng")
        if lat is None or lng is None:
            stats["skipped_no_coord"] += 1
            continue

        key = _coord_key(lat, lng)
        if key and key in coord_idx:
            cur.execute(
                """
                UPDATE pontos_ilp
                SET
                  bairro = COALESCE(bairro, ?),
                  endereco = CASE WHEN endereco IS NULL OR TRIM(endereco) = '' OR endereco = 'Sem endereco' THEN ? ELSE endereco END,
                  tipo_lampada = COALESCE(tipo_lampada, ?),
                  potencia = COALESCE(potencia, ?)
                WHERE id = ?
                """,
                (
                    p.get("bairro"),
                    p.get("endereco") or "Sem endereco",
                    p.get("tipo_lampada"),
                    p.get("potencia"),
                    coord_idx[key],
                ),
            )
            if cur.rowcount > 0:
                stats["updated"] += 1
            stats["dedup_coord"] += 1
            continue

        seq += 1
        etiqueta = normalize_text(p.get("etiqueta")) or f"IMPORT-{seq:06d}"
        base_tag = etiqueta
        n = 1
        while etiqueta in existing_tags:
            n += 1
            etiqueta = f"{base_tag}-{n}"
        existing_tags.add(etiqueta)

        cur.execute(
            """
            INSERT INTO pontos_ilp (
              etiqueta, endereco, bairro, cidade, lat, lng,
              tipo_lampada, potencia, status, origem_dado, confianca_dado
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                etiqueta,
                p.get("endereco") or "Sem endereco",
                p.get("bairro"),
                p.get("cidade") or "Macapa",
                lat,
                lng,
                p.get("tipo_lampada"),
                p.get("potencia"),
                p.get("status") or "cadastrado",
                origem_dado,
                0.85,
            ),
        )
        if key:
            coord_idx[key] = cur.lastrowid
        stats["inserted"] += 1

    return stats
